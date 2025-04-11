#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import pandas as pd
import numpy as np
import math
import argparse
import time
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
import psutil
import gc

# Thiết lập logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('excel_splitter.log')
    ]
)
logger = logging.getLogger(__name__)

def get_memory_usage():
    """Trả về memory usage hiện tại (MB)"""
    process = psutil.Process(os.getpid())
    return process.memory_info().rss / 1024 / 1024

def split_excel_file_with_optimized_memory(input_file, output_dir, sheets_per_file=3, rows_per_sheet=40000, 
                                          chunk_size=10000):
    """
    Chia file Excel lớn thành nhiều file nhỏ với tối ưu bộ nhớ
    
    Tham số:
        input_file (str): Đường dẫn đến file Excel đầu vào
        output_dir (str): Thư mục lưu các file đầu ra
        sheets_per_file (int): Số lượng sheet trong mỗi file đầu ra
        rows_per_sheet (int): Số hàng dữ liệu tối đa trong mỗi sheet
        chunk_size (int): Số lượng hàng đọc mỗi lần để tiết kiệm bộ nhớ
    """
    start_time = time.time()
    logger.info(f"Bắt đầu xử lý file: {input_file}")
    logger.info(f"Memory trước khi xử lý: {get_memory_usage():.2f} MB")
    
    # Tạo thư mục đầu ra nếu chưa tồn tại
    os.makedirs(output_dir, exist_ok=True)
    
    # Đọc danh sách sheets từ file đầu vào
    logger.info("Đọc danh sách sheets từ file đầu vào...")
    xl = pd.ExcelFile(input_file)
    sheet_names = xl.sheet_names
    
    logger.info(f"Tìm thấy {len(sheet_names)} sheets trong file đầu vào")
    
    # Tính tổng số hàng trong tất cả các sheets
    total_rows = 0
    rows_per_sheet_input = {}
    
    for sheet_name in sheet_names:
        # Chỉ đọc một hàng để lấy cấu trúc
        temp_df = xl.parse(sheet_name, nrows=1)
        columns = temp_df.columns
        
        # Sử dụng openpyxl để đếm số hàng mà không đọc hết dữ liệu
        wb = load_workbook(input_file, read_only=True)
        ws = wb[sheet_name]
        row_count = ws.max_row - 1  # Trừ đi header
        wb.close()
        
        rows_per_sheet_input[sheet_name] = (row_count, columns)
        total_rows += row_count
        logger.info(f"Sheet '{sheet_name}' có {row_count} hàng")
    
    # Tính toán số lượng file và sheet đầu ra
    total_output_sheets = math.ceil(total_rows / rows_per_sheet)
    total_output_files = math.ceil(total_output_sheets / sheets_per_file)
    
    logger.info(f"Tổng số hàng cần xử lý: {total_rows}")
    logger.info(f"Số sheets đầu ra dự kiến: {total_output_sheets}")
    logger.info(f"Số files đầu ra dự kiến: {total_output_files}")
    
    # Biến theo dõi quá trình
    rows_processed = 0
    current_output_file = 1
    current_sheet_in_file = 1
    current_rows_in_sheet = 0
    
    # Tạo file đầu ra đầu tiên
    output_filename = f"output_{current_output_file}.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    logger.info(f"Tạo file đầu ra mới: {output_path}")
    
    # Xử lý từng sheet trong file đầu vào
    for sheet_name in sheet_names:
        row_count, columns = rows_per_sheet_input[sheet_name]
        
        # Xử lý sheet theo từng chunk để tiết kiệm bộ nhớ
        for chunk_start in range(0, row_count, chunk_size):
            chunk_end = min(chunk_start + chunk_size, row_count)
            chunk_size_actual = chunk_end - chunk_start
            
            # Đọc một phần dữ liệu từ sheet
            logger.info(f"Đọc chunk {chunk_start}-{chunk_end} từ sheet '{sheet_name}'")
            df_chunk = xl.parse(sheet_name, skiprows=range(1, chunk_start + 1), nrows=chunk_size_actual)
            
            # Xử lý chunk
            chunk_rows_left = len(df_chunk)
            chunk_start_idx = 0
            
            while chunk_rows_left > 0:
                # Tính số hàng có thể thêm vào sheet hiện tại
                rows_to_add = min(chunk_rows_left, rows_per_sheet - current_rows_in_sheet)
                
                # Lấy phần dữ liệu cần thêm vào
                df_to_add = df_chunk.iloc[chunk_start_idx:chunk_start_idx + rows_to_add]
                
                # Tên của sheet đầu ra
                output_sheet_name = f"Sheet_{current_sheet_in_file}"
                
                # Thêm vào sheet hiện tại hoặc tạo mới
                if current_rows_in_sheet == 0:
                    # Sheet mới
                    logger.info(f"Tạo sheet mới '{output_sheet_name}' trong file {output_filename}")
                    df_to_add.to_excel(writer, sheet_name=output_sheet_name, index=False)
                else:
                    # Thêm vào sheet hiện tại
                    logger.info(f"Thêm {rows_to_add} hàng vào sheet '{output_sheet_name}' trong file {output_filename}")
                    
                    # Đọc dữ liệu hiện có
                    book = writer.book
                    sheet = book[output_sheet_name]
                    
                    # Thêm dữ liệu mới vào cuối, bỏ qua header
                    for r_idx, row in enumerate(dataframe_to_rows(df_to_add, index=False, header=False)):
                        for c_idx, value in enumerate(row):
                            sheet.cell(row=current_rows_in_sheet + r_idx + 2, column=c_idx + 1, value=value)
                
                # Cập nhật các biến đếm
                chunk_start_idx += rows_to_add
                chunk_rows_left -= rows_to_add
                current_rows_in_sheet += rows_to_add
                rows_processed += rows_to_add
                
                # Kiểm tra nếu sheet đã đầy
                if current_rows_in_sheet >= rows_per_sheet:
                    current_sheet_in_file += 1
                    current_rows_in_sheet = 0
                    
                    # Kiểm tra nếu file hiện tại đã đủ số sheet
                    if (current_sheet_in_file > sheets_per_file) and chunk_rows_left > 0:
                        # Lưu file hiện tại
                        logger.info(f"Lưu file: {output_path}")
                        writer.close()
                        
                        # Thu hồi bộ nhớ
                        gc.collect()
                        logger.info(f"Memory sau khi đóng file: {get_memory_usage():.2f} MB")
                        
                        # Tạo file mới
                        current_output_file += 1
                        current_sheet_in_file = 1
                        output_filename = f"output_{current_output_file}.xlsx"
                        output_path = os.path.join(output_dir, output_filename)
                        writer = pd.ExcelWriter(output_path, engine='openpyxl')
                        logger.info(f"Tạo file đầu ra mới: {output_path}")
            
            # Thu hồi bộ nhớ sau mỗi chunk
            del df_chunk
            gc.collect()
            logger.info(f"Memory sau khi xử lý chunk: {get_memory_usage():.2f} MB")
    
    # Đóng file đầu ra cuối cùng
    if writer is not None:
        logger.info(f"Lưu file đầu ra cuối cùng: {output_path}")
        writer.close()
    
    end_time = time.time()
    elapsed_time = end_time - start_time
    logger.info(f"Hoàn thành xử lý {total_rows} hàng trong {elapsed_time:.2f} giây")
    logger.info(f"Tốc độ xử lý: {total_rows / elapsed_time:.2f} hàng/giây")
    logger.info(f"Memory sau khi hoàn thành: {get_memory_usage():.2f} MB")
    
    return total_output_files

def main():
    parser = argparse.ArgumentParser(description='Chia file Excel lớn thành nhiều file nhỏ hơn')
    parser.add_argument('input_file', help='Đường dẫn đến file Excel đầu vào')
    parser.add_argument('output_dir', help='Thư mục lưu các file đầu ra')
    parser.add_argument('--sheets', type=int, default=3, help='Số lượng sheet trong mỗi file đầu ra (mặc định: 3)')
    parser.add_argument('--rows', type=int, default=40000, help='Số hàng dữ liệu tối đa trong mỗi sheet (mặc định: 40000)')
    parser.add_argument('--chunk-size', type=int, default=10000, help='Số hàng đọc mỗi lần để tiết kiệm bộ nhớ (mặc định: 10000)')
    
    args = parser.parse_args()
    
    # Kiểm tra file đầu vào tồn tại
    if not os.path.isfile(args.input_file):
        logger.error(f"Lỗi: File đầu vào không tồn tại: {args.input_file}")
        return
    
    try:
        num_files = split_excel_file_with_optimized_memory(
            args.input_file, 
            args.output_dir, 
            args.sheets, 
            args.rows,
            args.chunk_size
        )
        logger.info(f"Đã tạo thành công {num_files} file trong thư mục {args.output_dir}")
    except Exception as e:
        logger.exception(f"Lỗi khi xử lý file: {e}")

if __name__ == "__main__":
    main() 