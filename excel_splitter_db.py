#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import pandas as pd
import numpy as np
import math
import argparse
import time
import sqlite3
import multiprocessing as mp
from concurrent.futures import ProcessPoolExecutor, as_completed, ThreadPoolExecutor
import logging
import psutil
import gc
import threading
import tempfile
import shutil
from openpyxl import load_workbook

# Thiết lập logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('excel_splitter_db.log')
    ]
)
logger = logging.getLogger(__name__)

def get_memory_usage():
    """Trả về memory usage hiện tại (MB)"""
    process = psutil.Process(os.getpid())
    return process.memory_info().rss / 1024 / 1024

def create_database(db_path):
    """
    Tạo database SQLite với cấu trúc cần thiết
    
    Tham số:
        db_path (str): Đường dẫn đến file database
    
    Trả về:
        Kết nối đến database
    """
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Tạo bảng chính để lưu dữ liệu
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS excel_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        source_sheet INTEGER,
        row_num INTEGER,
        serial TEXT,
        qri TEXT
    )
    ''')
    
    # Tạo index để tối ưu truy vấn
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_source_sheet ON excel_data(source_sheet, row_num)')
    
    # Tạo bảng metadata để lưu thông tin
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS metadata (
        key TEXT PRIMARY KEY,
        value TEXT
    )
    ''')
    
    conn.commit()
    return conn

def read_sheet_to_db(input_file, sheet_name, sheet_idx, db_path):
    """
    Đọc dữ liệu từ một sheet trong file Excel và lưu vào database
    
    Tham số:
        input_file (str): Đường dẫn đến file Excel
        sheet_name (str): Tên sheet cần đọc
        sheet_idx (int): Chỉ số của sheet (1-based)
        db_path (str): Đường dẫn đến file database
    
    Trả về:
        Số lượng hàng đã đọc
    """
    try:
        start_time = time.time()
        logger.info(f"Đọc sheet {sheet_name} (#{sheet_idx})")
        
        # Đọc dữ liệu từ sheet
        df = pd.read_excel(input_file, sheet_name=sheet_name)
        total_rows = len(df)
        
        # Lưu dữ liệu vào database
        conn = sqlite3.connect(db_path)
        # Đảm bảo chỉ có 2 cột và đặt tên đúng
        if len(df.columns) >= 2:
            # Lấy hai cột đầu tiên và đổi tên
            df_subset = df.iloc[:, 0:2].copy()
            df_subset.columns = ['serial', 'qri']
            
            # Thêm thông tin sheet và số thứ tự hàng
            df_subset['source_sheet'] = sheet_idx
            df_subset['row_num'] = np.arange(1, len(df_subset) + 1)
            
            # Lưu vào database (bỏ qua header)
            df_subset.to_sql('excel_data', conn, if_exists='append', index=False)
        else:
            logger.warning(f"Sheet {sheet_name} không có đủ 2 cột")
        
        conn.commit()
        conn.close()
        
        elapsed_time = time.time() - start_time
        logger.info(f"Hoàn thành đọc sheet {sheet_name}: {total_rows} hàng trong {elapsed_time:.2f} giây")
        return total_rows - 1  # Trừ đi header
    
    except Exception as e:
        logger.error(f"Lỗi khi đọc sheet {sheet_name}: {e}")
        return 0

def count_data_rows(db_path):
    """
    Đếm tổng số hàng dữ liệu trong database
    
    Tham số:
        db_path (str): Đường dẫn đến file database
    
    Trả về:
        Tổng số hàng dữ liệu
    """
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM excel_data")
    count = cursor.fetchone()[0]
    conn.close()
    return count

def write_excel_file(output_path, sheet_idx, data_df, is_new_file=False):
    """
    Ghi dữ liệu vào file Excel đầu ra
    
    Tham số:
        output_path (str): Đường dẫn đến file Excel đầu ra
        sheet_idx (int): Chỉ số của sheet (1-based)
        data_df (DataFrame): Dữ liệu cần ghi
        is_new_file (bool): Có phải file mới không
    """
    sheet_name = f"CA {sheet_idx}"
    
    # Sử dụng lock để đồng bộ hóa truy cập vào cùng một file
    # Tạo thư mục cha nếu cần
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    
    try:
        # Xử lý file mới
        if is_new_file or not os.path.exists(output_path):
            # Tạo file mới với sheet đầu tiên
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
                data_df.to_excel(writer, sheet_name=sheet_name, index=False)
            return
        
        # Trường hợp file đã tồn tại
        # Đọc workbook hiện có
        if os.path.exists(output_path):
            # Kiểm tra nếu file có sẵn và có thể đọc được
            try:
                book = load_workbook(output_path)
                
                # Nếu sheet đã tồn tại, xóa nó để tránh lỗi
                if sheet_name in book.sheetnames:
                    idx = book.sheetnames.index(sheet_name)
                    book.remove(book.worksheets[idx])
                
                # Lưu workbook
                book.save(output_path)
                book.close()
                
                # Thêm sheet mới vào file
                with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
                    data_df.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception as e:
                # Nếu file bị hỏng, tạo file mới
                logger.warning(f"File đầu ra có thể bị hỏng, tạo lại: {e}")
                with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
                    data_df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # File không tồn tại, tạo mới
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
                data_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    except Exception as e:
        logger.error(f"Lỗi khi ghi file Excel {output_path}, sheet {sheet_name}: {e}")
        # Thử phương pháp khác, tạo file mới nếu ghi lỗi
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
                data_df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as inner_e:
            logger.error(f"Vẫn không thể ghi file: {inner_e}")

def get_data_for_sheet(db_path, start_row, row_count):
    """
    Lấy dữ liệu từ database cho một sheet đầu ra
    
    Tham số:
        db_path (str): Đường dẫn đến file database
        start_row (int): Hàng bắt đầu (1-based)
        row_count (int): Số lượng hàng cần lấy
    
    Trả về:
        DataFrame chứa dữ liệu
    """
    conn = sqlite3.connect(db_path)
    
    # Truy vấn dữ liệu
    query = f"""
    SELECT serial, qri FROM excel_data
    ORDER BY source_sheet, row_num
    LIMIT {row_count} OFFSET {start_row - 1}
    """
    
    df = pd.read_sql_query(query, conn)
    conn.close()
    
    return df

def create_output_files(db_path, output_dir, sheets_per_file, rows_per_sheet, max_workers):
    """
    Tạo các file Excel đầu ra từ dữ liệu trong database
    
    Tham số:
        db_path (str): Đường dẫn đến file database
        output_dir (str): Thư mục đầu ra
        sheets_per_file (int): Số lượng sheet trong mỗi file đầu ra
        rows_per_sheet (int): Số lượng hàng dữ liệu trong mỗi sheet đầu ra
        max_workers (int): Số lượng worker tối đa
    
    Trả về:
        Số lượng file đã tạo
    """
    # Đếm tổng số hàng dữ liệu
    total_rows = count_data_rows(db_path)
    logger.info(f"Tổng số hàng dữ liệu: {total_rows}")
    
    # Tính toán số lượng sheet và file cần tạo
    total_sheets = math.ceil(total_rows / rows_per_sheet)
    total_files = math.ceil(total_sheets / sheets_per_file)
    
    logger.info(f"Sẽ tạo {total_files} file, tổng cộng {total_sheets} sheet")
    
    # Tạo thư mục đầu ra nếu chưa tồn tại
    os.makedirs(output_dir, exist_ok=True)
    
    # Tạo danh sách các nhiệm vụ cần thực hiện
    tasks = []
    for file_idx in range(1, total_files + 1):
        for sheet_idx_in_file in range(1, sheets_per_file + 1):
            # Kiểm tra xem có phải sheet cuối cùng không
            global_sheet_idx = (file_idx - 1) * sheets_per_file + sheet_idx_in_file
            if global_sheet_idx <= total_sheets:
                tasks.append((file_idx, sheet_idx_in_file))
    
    # Tạo khóa cho mỗi file để tránh xung đột khi ghi
    file_locks = {}
    for file_idx in range(1, total_files + 1):
        output_filename = f"output_{file_idx}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        file_locks[output_path] = threading.RLock()
    
    def process_sheet(file_idx, sheet_idx_in_file):
        try:
            # Tính vị trí của sheet trong toàn bộ dữ liệu
            global_sheet_idx = (file_idx - 1) * sheets_per_file + sheet_idx_in_file
            
            # Tính vị trí bắt đầu và số lượng hàng cho sheet này
            start_row = (global_sheet_idx - 1) * rows_per_sheet + 1
            
            # Số hàng cần lấy cho sheet này
            sheet_row_count = min(rows_per_sheet, total_rows - (start_row - 1))
            
            if sheet_row_count <= 0:
                return False  # Không còn dữ liệu
            
            # Lấy dữ liệu từ database
            data_df = get_data_for_sheet(db_path, start_row, sheet_row_count)
            
            if data_df.empty:
                logger.warning(f"Không có dữ liệu cho file {file_idx}, sheet {sheet_idx_in_file}")
                return False
            
            # Đường dẫn file đầu ra
            output_filename = f"output_{file_idx}.xlsx"
            output_path = os.path.join(output_dir, output_filename)
            
            # Sử dụng lock cho file cụ thể
            with file_locks[output_path]:
                # Ghi dữ liệu vào file Excel
                is_new_file = sheet_idx_in_file == 1 and not os.path.exists(output_path)
                write_excel_file(output_path, sheet_idx_in_file, data_df, is_new_file)
            
            logger.info(f"Đã ghi {len(data_df)} hàng vào file {output_filename}, sheet {sheet_idx_in_file}")
            return True
        
        except Exception as e:
            logger.error(f"Lỗi khi tạo file {file_idx}, sheet {sheet_idx_in_file}: {e}")
            return False
    
    # Xử lý các sheet theo thứ tự file/sheet để tránh xung đột
    # Xử lý tuần tự từng file để tránh lỗi
    for file_idx in range(1, total_files + 1):
        # Xử lý tất cả sheet trong cùng một file
        for sheet_idx_in_file in range(1, sheets_per_file + 1):
            global_sheet_idx = (file_idx - 1) * sheets_per_file + sheet_idx_in_file
            if global_sheet_idx <= total_sheets:
                process_sheet(file_idx, sheet_idx_in_file)
    
    return total_files

def read_excel_to_db(input_file, db_path, total_sheets=None, max_workers=None):
    """
    Đọc dữ liệu từ file Excel và lưu vào database
    
    Tham số:
        input_file (str): Đường dẫn đến file Excel đầu vào
        db_path (str): Đường dẫn đến file database
        total_sheets (int): Tổng số sheet trong file đầu vào (nếu biết trước)
        max_workers (int): Số lượng worker tối đa
    
    Trả về:
        Tổng số hàng dữ liệu đã đọc
    """
    start_time = time.time()
    
    # Đọc danh sách sheet từ file Excel
    xl = pd.ExcelFile(input_file)
    sheet_names = xl.sheet_names
    
    # Nếu đã biết trước số sheet, kiểm tra xem có khớp không
    if total_sheets is not None and total_sheets != len(sheet_names):
        logger.warning(f"Cảnh báo: Số sheet thực tế ({len(sheet_names)}) khác với số sheet được chỉ định ({total_sheets})")
    
    logger.info(f"Đọc {len(sheet_names)} sheet từ file {input_file}")
    
    # Khởi tạo SQLite với WAL mode để hỗ trợ đa luồng tốt hơn
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.close()
    
    # Sử dụng multiprocessing để xử lý song song
    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = []
        for idx, sheet_name in enumerate(sheet_names, 1):
            futures.append(
                executor.submit(read_sheet_to_db, input_file, sheet_name, idx, db_path)
            )
        
        # Đợi tất cả các sheet được xử lý
        total_data_rows = 0
        for future in as_completed(futures):
            try:
                sheet_rows = future.result()
                total_data_rows += sheet_rows
            except Exception as e:
                logger.error(f"Lỗi khi xử lý sheet: {e}")
    
    elapsed_time = time.time() - start_time
    logger.info(f"Hoàn thành đọc {len(sheet_names)} sheet, {total_data_rows} hàng dữ liệu trong {elapsed_time:.2f} giây")
    
    return total_data_rows

def split_excel_with_db(input_file, output_dir, sheets_per_file=3, rows_per_sheet=40000, 
                        known_sheets=None, known_rows_per_sheet=None, max_workers=None, 
                        custom_db_path=None, delete_db=False):
    """
    Chia file Excel lớn thành nhiều file nhỏ sử dụng database làm trung gian
    
    Tham số:
        input_file (str): Đường dẫn đến file Excel đầu vào
        output_dir (str): Thư mục lưu các file đầu ra
        sheets_per_file (int): Số lượng sheet trong mỗi file đầu ra
        rows_per_sheet (int): Số hàng dữ liệu tối đa trong mỗi sheet
        known_sheets (int): Số sheet trong file đầu vào (nếu biết trước)
        known_rows_per_sheet (int): Số hàng trong mỗi sheet đầu vào (nếu biết trước)
        max_workers (int): Số lượng worker tối đa
        custom_db_path (str): Đường dẫn đến database tùy chỉnh (nếu muốn tái sử dụng)
        delete_db (bool): Có xóa database sau khi hoàn thành không
    """
    start_time = time.time()
    logger.info(f"Bắt đầu xử lý file: {input_file}")
    logger.info(f"Memory trước khi xử lý: {get_memory_usage():.2f} MB")
    
    # Xác định số lượng worker tối đa
    cpu_count = mp.cpu_count()
    if max_workers is None:
        max_workers = max(1, cpu_count - 1)  # Để lại ít nhất 1 core cho hệ thống
    
    logger.info(f"Sử dụng tối đa {max_workers} worker(s)")
    
    # Tạo thư mục tạm để lưu database nếu không có custom_db_path
    temp_dir = None
    if custom_db_path is None:
        temp_dir = tempfile.mkdtemp()
        db_path = os.path.join(temp_dir, "excel_data.db")
        logger.info(f"Tạo database tạm thời: {db_path}")
    else:
        db_path = custom_db_path
        logger.info(f"Sử dụng database hiện có: {db_path}")
    
    try:
        # Kiểm tra database đã tồn tại chưa
        db_exists = os.path.exists(db_path)
        
        if not db_exists:
            # Tạo database mới
            logger.info(f"Tạo database mới: {db_path}")
            conn = create_database(db_path)
            conn.close()
            
            # Đọc dữ liệu từ file Excel vào database
            logger.info("Đọc dữ liệu từ file Excel vào database...")
            read_excel_to_db(input_file, db_path, known_sheets, max_workers)
        else:
            logger.info(f"Sử dụng database hiện có: {db_path}")
        
        # Đếm tổng số hàng dữ liệu
        total_rows = count_data_rows(db_path)
        logger.info(f"Tổng số hàng dữ liệu đã đọc: {total_rows}")
        
        # Tạo các file Excel đầu ra
        logger.info("Tạo các file Excel đầu ra...")
        total_files = create_output_files(db_path, output_dir, sheets_per_file, rows_per_sheet, max_workers)
        
        # Tổng kết
        elapsed_time = time.time() - start_time
        logger.info(f"Hoàn thành xử lý trong {elapsed_time:.2f} giây")
        logger.info(f"Đã tạo {total_files} file trong thư mục {output_dir}")
        logger.info(f"Memory sau khi hoàn thành: {get_memory_usage():.2f} MB")
        
        return total_files
    
    finally:
        # Xóa thư mục tạm và database nếu cần
        if temp_dir is not None and delete_db:
            try:
                shutil.rmtree(temp_dir)
                logger.info(f"Đã xóa database tạm thời")
            except Exception as e:
                logger.error(f"Lỗi khi xóa database tạm thời: {e}")
        elif custom_db_path is not None and delete_db:
            try:
                os.remove(custom_db_path)
                logger.info(f"Đã xóa database tùy chỉnh: {custom_db_path}")
            except Exception as e:
                logger.error(f"Lỗi khi xóa database tùy chỉnh: {e}")
        elif not delete_db:
            if custom_db_path is not None:
                logger.info(f"Giữ lại database tùy chỉnh để tái sử dụng: {custom_db_path}")
            elif temp_dir is not None:
                db_path = os.path.join(temp_dir, "excel_data.db")
                # Sao chép database từ thư mục tạm ra thư mục đầu ra để lưu lại
                saved_db_path = os.path.join(output_dir, "excel_data.db")
                try:
                    shutil.copy2(db_path, saved_db_path)
                    logger.info(f"Đã lưu database vào: {saved_db_path}")
                    # Xóa thư mục tạm sau khi sao chép
                    shutil.rmtree(temp_dir)
                except Exception as e:
                    logger.error(f"Lỗi khi lưu database: {e}")

def main():
    parser = argparse.ArgumentParser(description='Chia file Excel lớn thành nhiều file nhỏ hơn sử dụng database làm trung gian')
    parser.add_argument('input_file', help='Đường dẫn đến file Excel đầu vào')
    parser.add_argument('output_dir', help='Thư mục lưu các file đầu ra')
    parser.add_argument('--sheets', type=int, default=3, help='Số lượng sheet trong mỗi file đầu ra (mặc định: 3)')
    parser.add_argument('--rows', type=int, default=40000, help='Số hàng DỮ LIỆU (không bao gồm header) tối đa trong mỗi sheet (mặc định: 40000)')
    parser.add_argument('--workers', type=int, default=None, help='Số lượng worker tối đa (mặc định: số CPU - 1)')
    parser.add_argument('--known-sheets', type=int, default=None, help='Số sheet trong file đầu vào (nếu biết trước)')
    parser.add_argument('--known-rows', type=int, default=None, help='Số hàng trong mỗi sheet đầu vào, bao gồm header (nếu biết trước)')
    parser.add_argument('--db', help='Đường dẫn đến database để tái sử dụng (nếu đã tồn tại)')
    parser.add_argument('--delete-db', type=int, default=0, help='Xóa database sau khi hoàn thành (0: không xóa, 1: xóa)')
    
    args = parser.parse_args()
    
    # Kiểm tra file đầu vào tồn tại nếu không dùng database có sẵn
    if args.db is None and not os.path.isfile(args.input_file):
        logger.error(f"Lỗi: File đầu vào không tồn tại: {args.input_file}")
        return
    
    # Kiểm tra tham số rows
    if args.rows <= 0:
        logger.error(f"Lỗi: Số hàng dữ liệu mỗi sheet phải lớn hơn 0")
        return
    
    # Kiểm tra tham số delete-db
    delete_db = args.delete_db == 1
    
    try:
        num_files = split_excel_with_db(
            args.input_file, 
            args.output_dir, 
            args.sheets, 
            args.rows,
            args.known_sheets,
            args.known_rows,
            args.workers,
            args.db,
            delete_db
        )
        logger.info(f"Đã tạo thành công {num_files} file trong thư mục {args.output_dir}")
    except Exception as e:
        logger.exception(f"Lỗi khi xử lý file: {e}")

if __name__ == "__main__":
    main() 
