#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import pandas as pd
import numpy as np
import math
import argparse
import time
import multiprocessing as mp
from concurrent.futures import ProcessPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
import psutil
import gc
import threading

# Thiết lập logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('excel_splitter_parallel.log')
    ]
)
logger = logging.getLogger(__name__)

# Khóa cho việc đồng bộ giữa các tiến trình
file_locks = {}
file_locks_lock = threading.Lock()

def get_memory_usage():
    """Trả về memory usage hiện tại (MB)"""
    process = psutil.Process(os.getpid())
    return process.memory_info().rss / 1024 / 1024

def process_sheet_chunk(input_file, sheet_name, chunk_start, chunk_size, skiprows=None):
    """
    Xử lý một chunk của sheet và trả về DataFrame
    
    Tham số:
        input_file (str): Đường dẫn đến file Excel đầu vào
        sheet_name (str): Tên sheet cần xử lý
        chunk_start (int): Vị trí bắt đầu của chunk
        chunk_size (int): Kích thước chunk
        skiprows (list): Các hàng cần bỏ qua
    
    Trả về:
        DataFrame chứa dữ liệu của chunk
    """
    # Nếu skiprows không được cung cấp, tính toán dựa trên chunk_start
    if skiprows is None and chunk_start > 0:
        skiprows = list(range(1, chunk_start + 1))
    
    # Đọc chunk từ file Excel
    df_chunk = pd.read_excel(input_file, sheet_name=sheet_name, skiprows=skiprows, nrows=chunk_size)
    
    return df_chunk

def get_file_lock(file_path):
    """Lấy lock cho file cụ thể để tránh race condition"""
    with file_locks_lock:
        if file_path not in file_locks:
            file_locks[file_path] = threading.RLock()
        return file_locks[file_path]

def write_data_to_sheet(writer, sheet_name, data, header=True, start_row=0):
    """
    Ghi dữ liệu vào sheet
    
    Tham số:
        writer (ExcelWriter): ExcelWriter đã được khởi tạo
        sheet_name (str): Tên sheet cần ghi
        data (DataFrame): Dữ liệu cần ghi
        header (bool): Có ghi header hay không
        start_row (int): Hàng bắt đầu ghi
    """
    if sheet_name not in writer.sheets:
        # Sheet mới
        data.to_excel(writer, sheet_name=sheet_name, index=False, header=header, if_sheet_exists='replace')
    else:
        # Sheet đã tồn tại
        book = writer.book
        sheet = book[sheet_name]
        
        # Xác định vị trí bắt đầu ghi
        if start_row == 0:
            # Tìm hàng đầu tiên trống
            for i in range(1, sheet.max_row + 2):
                if sheet.cell(row=i, column=1).value is None:
                    start_row = i - 1
                    break
            
            # Nếu không tìm thấy hàng trống, bắt đầu từ hàng cuối cùng
            if start_row == 0:
                start_row = sheet.max_row
        
        # Ghi dữ liệu
        rows = dataframe_to_rows(data, index=False, header=False)
        
        for r_idx, row in enumerate(rows):
            for c_idx, value in enumerate(row):
                sheet.cell(row=start_row + r_idx + 1, column=c_idx + 1, value=value)
    
    return

def count_rows_in_sheet(input_file, sheet_name):
    """
    Đếm số hàng trong một sheet của file Excel
    
    Tham số:
        input_file (str): Đường dẫn đến file Excel
        sheet_name (str): Tên sheet cần đếm
    
    Trả về:
        Tổng số hàng (bao gồm header) và số hàng dữ liệu (không bao gồm header)
    """
    try:
        # Sử dụng openpyxl để đếm số hàng mà không đọc hết dữ liệu
        wb = load_workbook(input_file, read_only=True)
        ws = wb[sheet_name]
        total_rows = ws.max_row
        data_rows = total_rows - 1  # Trừ đi header
        wb.close()
        return total_rows, data_rows
    except Exception as e:
        logger.error(f"Lỗi khi đếm số hàng trong sheet {sheet_name}: {e}")
        # Thử phương pháp khác nếu openpyxl không hoạt động
        try:
            df = pd.read_excel(input_file, sheet_name=sheet_name, nrows=1)
            # Đọc với iterrows để đếm số hàng
            with pd.ExcelFile(input_file) as xls:
                df_iter = pd.read_excel(xls, sheet_name=sheet_name, iterator=True, chunksize=1000)
                data_rows = sum(len(chunk) for chunk in df_iter) - 1  # Trừ đi header
                total_rows = data_rows + 1
            return total_rows, data_rows
        except Exception as inner_e:
            logger.error(f"Không thể đếm số hàng: {inner_e}")
            return 0, 0

def process_sheet(input_file, sheet_name, output_dir, worker_id, sheets_per_file=3, 
                data_rows_per_sheet=40000, chunk_size=None):
    """
    Xử lý toàn bộ một sheet từ file đầu vào và phân phối dữ liệu vào các file đầu ra
    
    Tham số:
        input_file (str): Đường dẫn đến file Excel đầu vào
        sheet_name (str): Tên sheet cần xử lý
        output_dir (str): Thư mục lưu các file đầu ra
        worker_id (int): ID của worker để tránh xung đột
        sheets_per_file (int): Số lượng sheet trong mỗi file đầu ra
        data_rows_per_sheet (int): Số hàng dữ liệu tối đa trong mỗi sheet (không bao gồm header)
        chunk_size (int): Số lượng hàng đọc mỗi lần để tiết kiệm bộ nhớ
    
    Trả về:
        files_info cập nhật
    """
    logger.info(f"Worker {worker_id}: Bắt đầu xử lý sheet: {sheet_name}")
    
    # Đếm số hàng trong sheet
    total_rows, data_rows = count_rows_in_sheet(input_file, sheet_name)
    
    logger.info(f"Worker {worker_id}: Sheet '{sheet_name}' có {total_rows} hàng tổng cộng, {data_rows} hàng dữ liệu (không bao gồm header)")
    logger.info(f"Worker {worker_id}: Mỗi sheet đầu ra sẽ chứa tối đa {data_rows_per_sheet} hàng dữ liệu (không bao gồm header)")
    
    # Nếu chunk_size không được chỉ định, sử dụng toàn bộ sheet
    if chunk_size is None or chunk_size <= 0:
        chunk_size = total_rows  # Đọc toàn bộ sheet trong một lần
        logger.info(f"Worker {worker_id}: Đọc toàn bộ sheet '{sheet_name}' trong một lần (chunk_size = {chunk_size})")
    
    # Tính toán file và sheet đầu ra cho worker này
    # Mỗi worker sẽ có file đầu ra riêng để tránh xung đột
    total_files_needed = math.ceil(data_rows / (data_rows_per_sheet * sheets_per_file))
    files_info = {}
    
    # Xử lý sheet theo từng chunk để tiết kiệm bộ nhớ
    for chunk_start in range(0, data_rows, chunk_size):
        chunk_end = min(chunk_start + chunk_size, data_rows)
        chunk_size_actual = chunk_end - chunk_start
        
        # Đọc một phần dữ liệu từ sheet
        logger.info(f"Worker {worker_id}: Đọc chunk {chunk_start}-{chunk_end} từ sheet '{sheet_name}'")
        
        if chunk_start == 0:
            # Chunk đầu tiên - bao gồm header
            df_chunk = process_sheet_chunk(input_file, sheet_name, 0, chunk_size_actual)
        else:
            # Các chunk tiếp theo - bỏ qua header
            skiprows = list(range(1, chunk_start + 1))
            df_chunk = process_sheet_chunk(input_file, sheet_name, 0, chunk_size_actual, skiprows)
        
        # Xử lý chunk
        chunk_rows_left = len(df_chunk)
        chunk_start_idx = 0
        
        while chunk_rows_left > 0:
            # Tính file/sheet để ghi dữ liệu dựa trên vị trí hiện tại
            current_row_position = chunk_start + chunk_start_idx
            file_idx = (worker_id * total_files_needed) + (current_row_position // (data_rows_per_sheet * sheets_per_file)) + 1
            sheet_idx = (current_row_position % (data_rows_per_sheet * sheets_per_file)) // data_rows_per_sheet + 1
            
            # Kiểm tra nếu vượt quá số lượng sheet trong file
            if sheet_idx > sheets_per_file:
                file_idx += 1
                sheet_idx = 1
            
            # Tên của sheet đầu ra
            output_sheet_name = f"Sheet_{sheet_idx}"
            
            # Tên file đầu ra
            output_filename = f"output_{file_idx}.xlsx"
            output_path = os.path.join(output_dir, output_filename)
            
            # Kiểm tra và cập nhật files_info
            if file_idx not in files_info:
                files_info[file_idx] = {"sheets": {}}
            
            if sheet_idx not in files_info[file_idx]["sheets"]:
                files_info[file_idx]["sheets"][sheet_idx] = 0
            
            # Xác định số lượng hàng có thể thêm vào
            current_rows = files_info[file_idx]["sheets"][sheet_idx]
            rows_to_add = min(chunk_rows_left, data_rows_per_sheet - current_rows)
            
            # Lấy phần dữ liệu cần thêm vào
            df_to_add = df_chunk.iloc[chunk_start_idx:chunk_start_idx + rows_to_add]
            
            # Sử dụng lock để tránh race condition
            file_lock = get_file_lock(output_path)
            with file_lock:
                try:
                    # Tạo thư mục nếu chưa tồn tại
                    os.makedirs(os.path.dirname(output_path), exist_ok=True)
                    
                    # Kiểm tra nếu file tồn tại
                    file_exists = os.path.exists(output_path)
                    
                    # Nếu file chưa tồn tại hoặc rỗng, tạo mới
                    if not file_exists or os.path.getsize(output_path) == 0:
                        # Tạo file mới
                        logger.info(f"Worker {worker_id}: Tạo file mới: {output_path}")
                        with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
                            logger.info(f"Worker {worker_id}: Tạo sheet mới '{output_sheet_name}'")
                            df_to_add.to_excel(writer, sheet_name=output_sheet_name, index=False)
                    else:
                        # Cập nhật file hiện có
                        try:
                            with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                                if current_rows == 0:
                                    # Sheet mới trong file hiện có
                                    logger.info(f"Worker {worker_id}: Tạo sheet mới '{output_sheet_name}' trong file {output_filename}")
                                    df_to_add.to_excel(writer, sheet_name=output_sheet_name, index=False)
                                else:
                                    # Thêm vào sheet hiện tại
                                    logger.info(f"Worker {worker_id}: Thêm {rows_to_add} hàng vào sheet '{output_sheet_name}' trong file {output_filename}")
                                    write_data_to_sheet(writer, output_sheet_name, df_to_add, header=False, start_row=current_rows)
                        except Exception as e:
                            logger.error(f"Worker {worker_id}: Lỗi khi cập nhật file {output_path}: {e}")
                            # Thử lại với phương pháp khác
                            try:
                                temp_df = pd.read_excel(output_path, sheet_name=None)
                                with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
                                    # Ghi lại tất cả các sheet hiện có
                                    for sheet_name, df in temp_df.items():
                                        if sheet_name == output_sheet_name and current_rows > 0:
                                            # Nối dữ liệu mới vào cuối
                                            df = pd.concat([df, df_to_add], ignore_index=True)
                                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                                        
                                    # Thêm sheet mới nếu chưa tồn tại
                                    if output_sheet_name not in temp_df and current_rows == 0:
                                        df_to_add.to_excel(writer, sheet_name=output_sheet_name, index=False)
                            except Exception as inner_e:
                                logger.error(f"Worker {worker_id}: Không thể khôi phục file {output_path}: {inner_e}")
                                continue
                except Exception as e:
                    logger.error(f"Worker {worker_id}: Lỗi khi xử lý file {output_path}: {e}")
                    continue
            
            # Cập nhật biến đếm
            files_info[file_idx]["sheets"][sheet_idx] += rows_to_add
            chunk_start_idx += rows_to_add
            chunk_rows_left -= rows_to_add
        
        # Thu hồi bộ nhớ sau mỗi chunk
        del df_chunk
        gc.collect()
        logger.info(f"Worker {worker_id}: Memory sau khi xử lý chunk: {get_memory_usage():.2f} MB")
    
    return files_info

def split_excel_file_parallel(input_file, output_dir, sheets_per_file=3, data_rows_per_sheet=40000, 
                             chunk_size=None, max_workers=None):
    """
    Chia file Excel lớn thành nhiều file nhỏ với xử lý song song
    
    Tham số:
        input_file (str): Đường dẫn đến file Excel đầu vào
        output_dir (str): Thư mục lưu các file đầu ra
        sheets_per_file (int): Số lượng sheet trong mỗi file đầu ra
        data_rows_per_sheet (int): Số hàng dữ liệu tối đa trong mỗi sheet (không bao gồm header)
        chunk_size (int): Số lượng hàng đọc mỗi lần để tiết kiệm bộ nhớ (None = đọc cả sheet)
        max_workers (int): Số lượng worker tối đa cho xử lý song song
    """
    start_time = time.time()
    logger.info(f"Bắt đầu xử lý file: {input_file}")
    logger.info(f"Memory trước khi xử lý: {get_memory_usage():.2f} MB")
    logger.info(f"Cấu hình: {sheets_per_file} sheets/file, {data_rows_per_sheet} hàng dữ liệu/sheet + 1 hàng header")
    
    # Tự động phát hiện chunk size nếu không được chỉ định
    if chunk_size is None:
        # Đọc sheet đầu tiên để ước tính kích thước tốt nhất
        logger.info("Phát hiện chunk size tối ưu...")
        try:
            xl = pd.ExcelFile(input_file)
            first_sheet = xl.sheet_names[0]
            total_rows, _ = count_rows_in_sheet(input_file, first_sheet)
            chunk_size = total_rows  # Đọc toàn bộ sheet trong một lần
            logger.info(f"Sử dụng chunk_size = {chunk_size} (toàn bộ sheet)")
        except Exception as e:
            logger.error(f"Không thể phát hiện chunk size tối ưu: {e}")
            chunk_size = 10000
            logger.info(f"Sử dụng chunk_size mặc định = {chunk_size}")
    
    # Xác định số lượng CPU có sẵn
    cpu_count = mp.cpu_count()
    if max_workers is None:
        max_workers = max(1, cpu_count - 1)  # Để lại ít nhất 1 core cho hệ thống
    
    logger.info(f"Sử dụng {max_workers} worker(s) để xử lý song song")
    
    # Tạo thư mục đầu ra nếu chưa tồn tại
    os.makedirs(output_dir, exist_ok=True)
    
    # Đọc danh sách sheets từ file đầu vào
    logger.info("Đọc danh sách sheets từ file đầu vào...")
    xl = pd.ExcelFile(input_file)
    sheet_names = xl.sheet_names
    
    logger.info(f"Tìm thấy {len(sheet_names)} sheets trong file đầu vào")
    
    # Điều chỉnh số lượng worker dựa trên số lượng sheet
    max_workers = min(max_workers, len(sheet_names))
    
    # Khởi tạo biến lưu trữ thông tin về các file đầu ra
    # Sử dụng Manager để chia sẻ dữ liệu giữa các tiến trình
    manager = mp.Manager()
    all_files_info = manager.dict()
    
    # Xử lý song song các sheet
    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        # Danh sách các future
        futures = []
        
        # Chia các sheet cho từng worker
        for worker_id, sheet_name in enumerate(sheet_names):
            future = executor.submit(
                process_sheet, 
                input_file, 
                sheet_name, 
                output_dir, 
                worker_id,
                sheets_per_file, 
                data_rows_per_sheet, 
                chunk_size
            )
            futures.append(future)
        
        # Đợi tất cả các worker hoàn thành
        for future in as_completed(futures):
            try:
                result = future.result()
                # Cập nhật all_files_info với kết quả từ worker
                for file_id, file_info in result.items():
                    if file_id not in all_files_info:
                        all_files_info[file_id] = file_info
                    else:
                        # Kết hợp thông tin sheet
                        for sheet_id, rows in file_info["sheets"].items():
                            if sheet_id not in all_files_info[file_id]["sheets"]:
                                all_files_info[file_id]["sheets"][sheet_id] = rows
                            else:
                                all_files_info[file_id]["sheets"][sheet_id] += rows
            except Exception as e:
                logger.error(f"Lỗi khi xử lý: {e}")
    
    # Tổng kết
    end_time = time.time()
    elapsed_time = end_time - start_time
    logger.info(f"Hoàn thành xử lý trong {elapsed_time:.2f} giây")
    logger.info(f"Đã tạo {len(all_files_info)} file trong thư mục {output_dir}")
    logger.info(f"Memory sau khi hoàn thành: {get_memory_usage():.2f} MB")
    
    return len(all_files_info)

def main():
    parser = argparse.ArgumentParser(description='Chia file Excel lớn thành nhiều file nhỏ hơn với xử lý song song')
    parser.add_argument('input_file', help='Đường dẫn đến file Excel đầu vào')
    parser.add_argument('output_dir', help='Thư mục lưu các file đầu ra')
    parser.add_argument('--sheets', type=int, default=3, help='Số lượng sheet trong mỗi file đầu ra (mặc định: 3)')
    parser.add_argument('--rows', type=int, default=40000, help='Số hàng DỮ LIỆU (không bao gồm header) tối đa trong mỗi sheet (mặc định: 40000)')
    parser.add_argument('--chunk-size', type=int, default=None, help='Số hàng đọc mỗi lần để tiết kiệm bộ nhớ (mặc định: đọc toàn bộ sheet)')
    parser.add_argument('--workers', type=int, default=None, help='Số lượng worker tối đa cho xử lý song song (mặc định: số CPU - 1)')
    
    args = parser.parse_args()
    
    # Kiểm tra file đầu vào tồn tại
    if not os.path.isfile(args.input_file):
        logger.error(f"Lỗi: File đầu vào không tồn tại: {args.input_file}")
        return
    
    # Kiểm tra tham số rows
    if args.rows <= 0:
        logger.error(f"Lỗi: Số hàng dữ liệu mỗi sheet phải lớn hơn 0")
        return
    
    try:
        num_files = split_excel_file_parallel(
            args.input_file, 
            args.output_dir, 
            args.sheets, 
            args.rows,
            args.chunk_size,
            args.workers
        )
        logger.info(f"Đã tạo thành công {num_files} file trong thư mục {args.output_dir}")
    except Exception as e:
        logger.exception(f"Lỗi khi xử lý file: {e}")

if __name__ == "__main__":
    main() 