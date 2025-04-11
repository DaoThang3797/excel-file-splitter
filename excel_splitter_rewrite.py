#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import pandas as pd
import numpy as np
import math
import argparse
import time
import multiprocessing as mp
from concurrent.futures import ProcessPoolExecutor, as_completed, ThreadPoolExecutor
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
import psutil
import gc
import threading
import queue

# Thiết lập logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('excel_splitter_rewrite.log')
    ]
)
logger = logging.getLogger(__name__)

# Khóa cho việc đồng bộ giữa các tiến trình
file_locks = {}
file_locks_lock = threading.Lock()

def get_memory_usage():
    """Trả về memory usage hiện tại (MB)"""
    process = psutil.Process(os.getpid())
    return process.memory_info().rss / 1024 / 1024

def get_file_lock(file_path):
    """Lấy lock cho file cụ thể để tránh race condition"""
    with file_locks_lock:
        if file_path not in file_locks:
            file_locks[file_path] = threading.RLock()
        return file_locks[file_path]

def count_rows_in_sheet(input_file, sheet_name):
    """
    Đếm số hàng trong một sheet của file Excel
    
    Tham số:
        input_file (str): Đường dẫn đến file Excel
        sheet_name (str): Tên sheet cần đếm
    
    Trả về:
        Tổng số hàng (bao gồm header) và số hàng dữ liệu (không bao gồm header)
    """
    try:
        # Sử dụng openpyxl để đếm số hàng mà không đọc hết dữ liệu
        wb = load_workbook(input_file, read_only=True)
        ws = wb[sheet_name]
        total_rows = ws.max_row
        data_rows = total_rows - 1  # Trừ đi header
        wb.close()
        return total_rows, data_rows
    except Exception as e:
        logger.error(f"Lỗi khi đếm số hàng trong sheet {sheet_name}: {e}")
        # Thử phương pháp khác nếu openpyxl không hoạt động
        try:
            df = pd.read_excel(input_file, sheet_name=sheet_name, nrows=1)
            # Đọc với iterrows để đếm số hàng
            with pd.ExcelFile(input_file) as xls:
                df_iter = pd.read_excel(xls, sheet_name=sheet_name, iterator=True, chunksize=1000)
                data_rows = sum(len(chunk) for chunk in df_iter) - 1  # Trừ đi header
                total_rows = data_rows + 1
            return total_rows, data_rows
        except Exception as inner_e:
            logger.error(f"Không thể đếm số hàng: {inner_e}")
            return 0, 0

def read_sheet_data(input_file, sheet_name, start_row=0, num_rows=None):
    """
    Đọc dữ liệu từ một sheet trong file Excel
    
    Tham số:
        input_file (str): Đường dẫn đến file Excel
        sheet_name (str): Tên của sheet cần đọc
        start_row (int): Hàng bắt đầu đọc (0 = header)
        num_rows (int): Số hàng cần đọc
    
    Trả về:
        DataFrame chứa dữ liệu được đọc
    """
    try:
        if start_row == 0:
            # Đọc bao gồm header
            df = pd.read_excel(input_file, sheet_name=sheet_name, nrows=num_rows)
        else:
            # Skip header và các hàng trước start_row
            skiprows = list(range(1, start_row + 1))
            df = pd.read_excel(input_file, sheet_name=sheet_name, skiprows=skiprows, nrows=num_rows)
            
            # Đọc header từ file gốc để giữ nguyên cấu trúc
            header_df = pd.read_excel(input_file, sheet_name=sheet_name, nrows=1)
            df.columns = header_df.columns
        
        return df
    except Exception as e:
        logger.error(f"Lỗi khi đọc dữ liệu từ sheet {sheet_name}: {e}")
        return pd.DataFrame()

def analyze_input_file(input_file):
    """
    Phân tích file đầu vào để xác định cấu trúc và tổng số dữ liệu
    
    Tham số:
        input_file (str): Đường dẫn đến file Excel đầu vào
    
    Trả về:
        Thông tin về số sheet, số hàng mỗi sheet, và tổng số hàng dữ liệu
    """
    logger.info(f"Phân tích file đầu vào: {input_file}")
    
    # Đọc thông tin sheet
    xl = pd.ExcelFile(input_file)
    sheet_names = xl.sheet_names
    
    sheets_info = {}
    total_data_rows = 0
    
    # Đọc thông tin từng sheet
    for sheet_name in sheet_names:
        total_rows, data_rows = count_rows_in_sheet(input_file, sheet_name)
        sheets_info[sheet_name] = {
            "total_rows": total_rows,
            "data_rows": data_rows
        }
        total_data_rows += data_rows
    
    result = {
        "num_sheets": len(sheet_names),
        "sheet_names": sheet_names,
        "sheets_info": sheets_info,
        "total_data_rows": total_data_rows
    }
    
    logger.info(f"Phân tích hoàn tất: {len(sheet_names)} sheets, {total_data_rows} hàng dữ liệu tổng cộng")
    return result

def calculate_output_distribution(input_analysis, sheets_per_file=3, data_rows_per_sheet=40000):
    """
    Tính toán cách phân phối dữ liệu vào các file đầu ra
    
    Tham số:
        input_analysis (dict): Kết quả phân tích file đầu vào
        sheets_per_file (int): Số lượng sheet trong mỗi file đầu ra
        data_rows_per_sheet (int): Số hàng dữ liệu (không bao gồm header) trong mỗi sheet đầu ra
    
    Trả về:
        Thông tin về cách phân phối dữ liệu vào các file/sheet đầu ra
    """
    total_data_rows = input_analysis["total_data_rows"]
    
    # Tính tổng số sheet đầu ra cần thiết
    total_output_sheets = math.ceil(total_data_rows / data_rows_per_sheet)
    
    # Tính tổng số file đầu ra cần thiết
    total_output_files = math.ceil(total_output_sheets / sheets_per_file)
    
    logger.info(f"Phân phối dữ liệu: {total_output_sheets} sheets đầu ra, {total_output_files} files đầu ra")
    
    # Tạo cấu trúc mô tả file đầu ra
    output_distribution = {
        "files": {},
        "total_output_files": total_output_files,
        "total_output_sheets": total_output_sheets
    }
    
    # Tính toán phân phối dữ liệu vào từng file/sheet
    remaining_data_rows = total_data_rows
    current_file = 1
    current_sheet = 1
    
    # Tạo danh sách các phần dữ liệu cần đọc
    data_segments = []
    
    # Theo dõi vị trí hiện tại trong từng sheet đầu vào
    current_positions = {sheet_name: 0 for sheet_name in input_analysis["sheet_names"]}
    
    # Xử lý lần lượt các sheet đầu vào
    for sheet_name in input_analysis["sheet_names"]:
        sheet_data_rows = input_analysis["sheets_info"][sheet_name]["data_rows"]
        rows_processed = 0
        
        while rows_processed < sheet_data_rows:
            # Số hàng còn lại trong sheet hiện tại
            rows_left_in_sheet = sheet_data_rows - rows_processed
            
            # Số hàng còn trống trong sheet đầu ra hiện tại
            if current_file not in output_distribution["files"]:
                output_distribution["files"][current_file] = {"sheets": {}}
            
            if current_sheet not in output_distribution["files"][current_file]["sheets"]:
                output_distribution["files"][current_file]["sheets"][current_sheet] = 0
            
            current_output_rows = output_distribution["files"][current_file]["sheets"][current_sheet]
            rows_available_in_output = data_rows_per_sheet - current_output_rows
            
            # Số hàng sẽ được sao chép
            rows_to_copy = min(rows_left_in_sheet, rows_available_in_output)
            
            if rows_to_copy > 0:
                # Thêm phân đoạn dữ liệu để đọc sau này
                start_row = current_positions[sheet_name]
                data_segments.append({
                    "input_sheet": sheet_name,
                    "start_row": start_row,
                    "num_rows": rows_to_copy,
                    "output_file": current_file,
                    "output_sheet": current_sheet
                })
                
                # Cập nhật vị trí hiện tại trong sheet đầu vào
                current_positions[sheet_name] += rows_to_copy
                
                # Cập nhật trạng thái đầu ra
                output_distribution["files"][current_file]["sheets"][current_sheet] += rows_to_copy
                
                # Cập nhật số hàng đã xử lý
                rows_processed += rows_to_copy
                remaining_data_rows -= rows_to_copy
                
                # Kiểm tra nếu sheet đầu ra đã đầy
                if output_distribution["files"][current_file]["sheets"][current_sheet] >= data_rows_per_sheet:
                    current_sheet += 1
                    
                    # Kiểm tra nếu file đầu ra đã đủ số sheet
                    if current_sheet > sheets_per_file:
                        current_file += 1
                        current_sheet = 1
            else:
                # Không còn chỗ trống trong sheet hiện tại, chuyển sang sheet tiếp theo
                current_sheet += 1
                if current_sheet > sheets_per_file:
                    current_file += 1
                    current_sheet = 1
    
    output_distribution["data_segments"] = data_segments
    return output_distribution

def process_data_segment(input_file, segment, output_dir):
    """
    Xử lý một phân đoạn dữ liệu: đọc từ file đầu vào và ghi vào file đầu ra
    
    Tham số:
        input_file (str): Đường dẫn đến file Excel đầu vào
        segment (dict): Thông tin về phân đoạn dữ liệu cần xử lý
        output_dir (str): Thư mục lưu các file đầu ra
    """
    try:
        input_sheet = segment["input_sheet"]
        start_row = segment["start_row"]
        num_rows = segment["num_rows"]
        output_file_idx = segment["output_file"]
        output_sheet_idx = segment["output_sheet"]
        
        # Tên file đầu ra
        output_filename = f"output_{output_file_idx}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        
        # Tên sheet đầu ra
        output_sheet_name = f"Sheet_{output_sheet_idx}"
        
        # Đọc dữ liệu từ file đầu vào
        include_header = start_row == 0
        df = read_sheet_data(input_file, input_sheet, start_row, num_rows)
        
        if df.empty:
            logger.error(f"Không thể đọc dữ liệu từ sheet {input_sheet} ở vị trí {start_row}")
            return
        
        # Sử dụng lock để tránh race condition
        file_lock = get_file_lock(output_path)
        with file_lock:
            # Kiểm tra xem file đã tồn tại chưa
            file_exists = os.path.exists(output_path)
            
            # Xác định sheet có tồn tại hay không
            sheet_exists = False
            if file_exists:
                try:
                    existing_sheets = pd.ExcelFile(output_path).sheet_names
                    sheet_exists = output_sheet_name in existing_sheets
                except:
                    file_exists = False  # File có thể bị hỏng
            
            # Chuẩn bị mode và thông số cho ExcelWriter
            mode = 'a' if file_exists else 'w'
            
            try:
                if not file_exists:
                    # Tạo thư mục đầu ra nếu chưa tồn tại
                    os.makedirs(os.path.dirname(output_path), exist_ok=True)
                    logger.info(f"Tạo file mới: {output_path}")
                
                with pd.ExcelWriter(output_path, engine='openpyxl', mode=mode, if_sheet_exists='overlay') as writer:
                    if not sheet_exists:
                        # Sheet mới
                        logger.info(f"Ghi {len(df)} hàng vào sheet mới '{output_sheet_name}' trong file {output_filename}")
                        df.to_excel(writer, sheet_name=output_sheet_name, index=False)
                    else:
                        # Thêm vào sheet hiện có
                        logger.info(f"Thêm {len(df)} hàng vào sheet '{output_sheet_name}' trong file {output_filename}")
                        
                        # Đọc sheet hiện tại
                        existing_data = pd.read_excel(output_path, sheet_name=output_sheet_name)
                        
                        # Nối dữ liệu mới, giữ lại header của dữ liệu hiện có
                        combined_data = pd.concat([existing_data, df.iloc[1:] if include_header else df], ignore_index=True)
                        
                        # Ghi đè sheet
                        combined_data.to_excel(writer, sheet_name=output_sheet_name, index=False)
            except Exception as e:
                logger.error(f"Lỗi khi ghi vào {output_path}, sheet {output_sheet_name}: {e}")
                # Thử phương pháp khác
                try:
                    # Đọc tất cả các sheet
                    all_sheets = {}
                    xl = pd.ExcelFile(output_path)
                    for sheet in xl.sheet_names:
                        all_sheets[sheet] = pd.read_excel(output_path, sheet_name=sheet)
                    
                    # Cập nhật hoặc thêm sheet mới
                    if output_sheet_name in all_sheets:
                        all_sheets[output_sheet_name] = pd.concat(
                            [all_sheets[output_sheet_name], df.iloc[1:] if include_header else df], 
                            ignore_index=True
                        )
                    else:
                        all_sheets[output_sheet_name] = df
                    
                    # Ghi lại toàn bộ file
                    with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
                        for sheet_name, sheet_data in all_sheets.items():
                            sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)
                except Exception as inner_e:
                    logger.error(f"Không thể khôi phục file {output_path}: {inner_e}")
    except Exception as e:
        logger.error(f"Lỗi khi xử lý phân đoạn dữ liệu: {e}")

def split_excel_file_distributed(input_file, output_dir, sheets_per_file=3, data_rows_per_sheet=40000, max_workers=None):
    """
    Chia file Excel lớn thành nhiều file nhỏ với cách tiếp cận phân phối dữ liệu mới
    
    Tham số:
        input_file (str): Đường dẫn đến file Excel đầu vào
        output_dir (str): Thư mục lưu các file đầu ra
        sheets_per_file (int): Số lượng sheet trong mỗi file đầu ra
        data_rows_per_sheet (int): Số hàng dữ liệu tối đa trong mỗi sheet (không bao gồm header)
        max_workers (int): Số lượng worker tối đa cho xử lý song song
    """
    start_time = time.time()
    logger.info(f"Bắt đầu xử lý file: {input_file}")
    logger.info(f"Memory trước khi xử lý: {get_memory_usage():.2f} MB")
    logger.info(f"Cấu hình: {sheets_per_file} sheets/file, {data_rows_per_sheet} hàng dữ liệu/sheet + 1 hàng header")
    
    # Tạo thư mục đầu ra nếu chưa tồn tại
    os.makedirs(output_dir, exist_ok=True)
    
    # Xác định số lượng worker tối đa
    cpu_count = mp.cpu_count()
    if max_workers is None:
        max_workers = max(1, cpu_count - 1)  # Để lại ít nhất 1 core cho hệ thống
    
    logger.info(f"Sử dụng tối đa {max_workers} worker(s) cho xử lý song song")
    
    # Phân tích file đầu vào
    input_analysis = analyze_input_file(input_file)
    
    # Tính toán phân phối dữ liệu
    distribution = calculate_output_distribution(input_analysis, sheets_per_file, data_rows_per_sheet)
    
    # Lấy danh sách các phân đoạn dữ liệu cần xử lý
    data_segments = distribution["data_segments"]
    
    # Tạo trước các file đầu ra trống
    logger.info(f"Tạo {distribution['total_output_files']} file đầu ra")
    
    # Xử lý các phân đoạn dữ liệu song song
    logger.info(f"Xử lý {len(data_segments)} phân đoạn dữ liệu")
    
    # Điều chỉnh số worker dựa trên số phân đoạn dữ liệu
    active_workers = min(max_workers, len(data_segments))
    
    with ThreadPoolExecutor(max_workers=active_workers) as executor:
        futures = []
        for segment in data_segments:
            futures.append(
                executor.submit(process_data_segment, input_file, segment, output_dir)
            )
        
        # Đợi tất cả các worker hoàn thành
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                logger.error(f"Lỗi khi xử lý phân đoạn dữ liệu: {e}")
    
    # Tổng kết
    end_time = time.time()
    elapsed_time = end_time - start_time
    logger.info(f"Hoàn thành xử lý trong {elapsed_time:.2f} giây")
    logger.info(f"Đã tạo {distribution['total_output_files']} file trong thư mục {output_dir}")
    logger.info(f"Memory sau khi hoàn thành: {get_memory_usage():.2f} MB")
    
    return distribution['total_output_files']

def main():
    parser = argparse.ArgumentParser(description='Chia file Excel lớn thành nhiều file nhỏ hơn với cách tiếp cận phân phối dữ liệu mới')
    parser.add_argument('input_file', help='Đường dẫn đến file Excel đầu vào')
    parser.add_argument('output_dir', help='Thư mục lưu các file đầu ra')
    parser.add_argument('--sheets', type=int, default=3, help='Số lượng sheet trong mỗi file đầu ra (mặc định: 3)')
    parser.add_argument('--rows', type=int, default=40000, help='Số hàng DỮ LIỆU (không bao gồm header) tối đa trong mỗi sheet (mặc định: 40000)')
    parser.add_argument('--workers', type=int, default=None, help='Số lượng worker tối đa cho xử lý song song (mặc định: số CPU - 1)')
    
    args = parser.parse_args()
    
    # Kiểm tra file đầu vào tồn tại
    if not os.path.isfile(args.input_file):
        logger.error(f"Lỗi: File đầu vào không tồn tại: {args.input_file}")
        return
    
    # Kiểm tra tham số rows
    if args.rows <= 0:
        logger.error(f"Lỗi: Số hàng dữ liệu mỗi sheet phải lớn hơn 0")
        return
    
    try:
        num_files = split_excel_file_distributed(
            args.input_file, 
            args.output_dir, 
            args.sheets, 
            args.rows,
            args.workers
        )
        logger.info(f"Đã tạo thành công {num_files} file trong thư mục {args.output_dir}")
    except Exception as e:
        logger.exception(f"Lỗi khi xử lý file: {e}")

if __name__ == "__main__":
    main() 